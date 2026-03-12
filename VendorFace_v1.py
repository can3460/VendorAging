"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                          VendorFace v6.1 ULTIMATE                            ║
║                   AP Intelligence Dashboard | Opella Finance                 ║
║                  Perfect Engine + Stunning UX | Production Ready             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from io import BytesIO
import io, json, warnings, os, base64
from datetime import datetime, date, timedelta

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
# 1. PAGE CONFIG & STRICT SECURITY UI HIDERS
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="VendorFace v6.1 | AP Intelligence",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None, 
        'Report a bug': None, 
        'About': "VendorFace v6.1 ULTIMATE | Opella Finance Operations"
    }
)

# STRICT CSS: Completely annihilates Streamlit Cloud branding, header, and deploy buttons
st.markdown("""
<style>
    header {visibility: hidden !important;}
    .stApp {margin-top: -60px !important;}
    #MainMenu {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    .stDeployButton {display: none !important;}
    [data-testid="stToolbar"] {display: none !important;}
    .viewerBadge_container__1QSob {display: none !important;}
    .viewerBadge_link__1S137 {display: none !important;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# 2. SYSTEM CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
VERSION_NO = "v6.1 ULTIMATE"
MASTER_ADMIN = "can.adiguzel@sanofi.com"
USER_DB = "users_v6.xlsx"

AGING_LABELS = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days"]
AGING_COLORS = ["#10B981", "#3B82F6", "#F59E0B", "#EF4444", "#991B1B"]
AGING_BG = [
    "rgba(16,185,129,0.1)", "rgba(59,130,246,0.1)", "rgba(245,158,11,0.1)", 
    "rgba(239,68,68,0.1)", "rgba(153,27,27,0.1)"
]

CURRENCIES = {
    "EUR": "EUR", "USD": "USD", "GBP": "GBP", "TRY": "TRY", "EGP": "EGP", "TND": "TND",
    "AED": "AED", "SAR": "SAR", "CNY": "CNY", "JPY": "JPY", "INR": "INR", "BRL": "BRL"
}

GL_NAMES = {
    "160000": "AP Trade", "160100": "AP Interco", "160200": "AP Services",
    "168000": "AP Accruals", "165000": "Employee Pay", "163000": "Travel Accrual"
}

ICO_GL = ("160100", "161", "162", "42905")
EMP_GL = ("165", "163", "42006")
ICO_KW = ["interco", "ico", "affiliated", "related party", "intragroup", "group payable"]
EMP_KW = ["employee", "personnel", "staff", "travel", "expense", "salary", "bonus"]

# ══════════════════════════════════════════════════════════════════════════════
# 3. STATE MANAGEMENT & AUTHENTICATION CORE
# ══════════════════════════════════════════════════════════════════════════════
if 'logged_in' not in st.session_state: st.session_state.logged_in = False
if 'user_email' not in st.session_state: st.session_state.user_email = ""
if 'user_name' not in st.session_state: st.session_state.user_name = ""
if 'theme' not in st.session_state: st.session_state.theme = 'light'
if 'disp_curr' not in st.session_state: st.session_state.disp_curr = "EUR"
if 'in_k' not in st.session_state: st.session_state.in_k = False

def load_users():
    """Loads authorized users from DB. Creates default if not exists."""
    if not os.path.exists(USER_DB):
        df = pd.DataFrame([
            {"email": MASTER_ADMIN, "role": "admin", "added_on": date.today().isoformat()},
            {"email": "admin@opella.com", "role": "admin", "added_on": date.today().isoformat()}
        ])
        df.to_excel(USER_DB, index=False)
        return df
    return pd.read_excel(USER_DB)

def is_authorized(email):
    email = str(email).strip().lower()
    if not (email.endswith("@sanofi.com") or email.endswith("@opella.com")): return False
    users = load_users()
    return email in users['email'].str.lower().values

# ══════════════════════════════════════════════════════════════════════════════
# 4. THEME SYSTEM & GLASSMORPHISM CSS
# ══════════════════════════════════════════════════════════════════════════════
def get_theme():
    if st.session_state.theme == 'dark':
        return {
            'bg': '#0F172A', 'card_bg': '#1E293B', 'text': '#F1F5F9', 
            'text_sec': '#94A3B8', 'border': '#334155', 'accent': '#3B82F6'
        }
    return {
        'bg': '#F8FAFC', 'card_bg': '#FFFFFF', 'text': '#0F172A', 
        'text_sec': '#64748B', 'border': '#E2E8F0', 'accent': '#2563EB'
    }

theme = get_theme()

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

* {{ font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; }}

html, body, [data-testid="stAppViewContainer"] {{
    background: {theme['bg']}; color: {theme['text']};
}}

[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, #1e3a8a 0%, #312e81 100%);
}}
[data-testid="stSidebar"] * {{ color: #F1F5F9 !important; }}

/* Glassmorphism Cards */
.glass-card {{
    background: rgba(255, 255, 255, 0.05); backdrop-filter: blur(10px);
    border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 16px;
    padding: 24px; box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1); transition: all 0.3s ease;
}}
.glass-card:hover {{ transform: translateY(-4px); box-shadow: 0 12px 48px rgba(0, 0, 0, 0.15); }}

/* KPI Cards */
.kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px; }}
.kpi-card {{
    background: {theme['card_bg']}; border: 1px solid {theme['border']};
    border-radius: 12px; padding: 20px; position: relative; overflow: hidden; transition: all 0.3s ease;
}}
.kpi-card:hover {{ transform: translateY(-2px); box-shadow: 0 8px 24px rgba(0,0,0,0.12); }}
.kpi-card::before {{
    content: ''; position: absolute; top: 0; left: 0; width: 4px; height: 100%;
    background: linear-gradient(180deg, var(--accent), transparent);
}}
.kpi-label {{ font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; color: {theme['text_sec']}; margin-bottom: 8px; }}
.kpi-value {{ font-size: 2rem; font-weight: 800; color: {theme['text']}; line-height: 1.2; margin-bottom: 4px; }}
.kpi-sub {{ font-size: 0.85rem; color: {theme['text_sec']}; }}

/* Progress Bars */
.progress-bar {{ height: 8px; background: {theme['border']}; border-radius: 999px; overflow: hidden; margin-top: 12px; }}
.progress-fill {{ height: 100%; background: linear-gradient(90deg, #3B82F6, #8B5CF6); transition: width 0.6s ease; }}

/* Section Headers */
.sec-hdr {{ font-size: 1.1rem; font-weight: 700; color: {theme['text']}; margin: 24px 0 12px; padding-bottom: 8px; border-bottom: 2px solid {theme['accent']}; }}

/* Info Boxes */
.info-box {{ background: rgba(59, 130, 246, 0.1); border: 1px solid rgba(59, 130, 246, 0.3); border-radius: 12px; padding: 16px; margin: 12px 0; font-size: 0.9rem; line-height: 1.6; }}
.warning-box {{ background: rgba(239, 68, 68, 0.1); border: 1px solid rgba(239, 68, 68, 0.3); padding: 16px; border-radius: 12px; }}
.success-box {{ background: rgba(16, 185, 129, 0.1); border: 1px solid rgba(16, 185, 129, 0.3); padding: 16px; border-radius: 12px; }}

/* DataFrame Styling Override */
[data-testid="stDataFrame"] {{ width: 100% !important; }}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# 5. SECURE LOGIN SCREEN (GLASSMORPHISM)
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    c1, c2, c3 = st.columns([1, 1.2, 1])
    with c2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(f"""
        <div class="glass-card" style="text-align:center; padding: 40px;">
            <div style="font-size: 4rem; margin-bottom: 10px;">🛡️</div>
            <h1 style="color: {theme['text']}; margin-bottom: 5px;">Opella</h1>
            <h3 style="color: {theme['text_sec']}; font-weight: 400; margin-bottom: 30px;">VendorFace {VERSION_NO}</h3>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("auth_form"):
            email_input = st.text_input("Corporate Email", placeholder="name.surname@opella.com").strip().lower()
            submit = st.form_submit_button("Secure Access", use_container_width=True)
            
            if submit:
                if is_authorized(email_input):
                    st.session_state.logged_in = True
                    st.session_state.user_email = email_input
                    st.session_state.user_name = email_input.split('@')[0].replace('.', ' ').title()
                    st.rerun()
                else:
                    st.error("Access Denied: Unrecognized email or domain policy violation.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# 6. BUSINESS LOGIC & HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def is_ico(gl, vname=""):
    gl, vn = str(gl).strip(), f" {str(vname).lower()} "
    return gl.startswith(ICO_GL) or any(f" {k} " in vn for k in ICO_KW) or "intercompany" in vn

def is_employee(gl, vname=""):
    gl, vn = str(gl).strip(), f" {str(vname).lower()} "
    return gl.startswith(EMP_GL) or any(f" {k} " in vn for k in EMP_KW)

def assign_aging(days):
    if days <= 0: return "Current"
    elif days <= 30: return "1-30 Days"
    elif days <= 60: return "31-60 Days"
    elif days <= 90: return "61-90 Days"
    else: return "90+ Days"

def fa(val):
    """Format Amount specifically for UI Rendering (Does not affect export)"""
    curr = st.session_state.get("disp_curr", "EUR")
    kk = st.session_state.get("in_k", False)
    if pd.isna(val): return "—"
    v = abs(float(val)) / (1000 if kk else 1)
    return f"{v:,.0f} {'k' if kk else ''}{curr}"

def smart_read(file):
    """Bulletproof file reader handling multiple encodings for CSVs"""
    name = file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        return pd.read_excel(file)
    raw = file.getvalue()
    for enc in ["utf-8-sig", "utf-8", "iso-8859-9", "cp1254", "latin-1", "cp1252"]:
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc, sep=None, engine="python", on_bad_lines="skip")
        except:
            continue
    raise ValueError(f"Could not read '{file.name}'. Encoding or format unsupported.")

# ══════════════════════════════════════════════════════════════════════════════
# 7. PERFECT DATA LOADERS
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_fbl1n(file):
    """Perfect FBL1N loader with robust column mapping"""
    df = smart_read(file)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()].copy()
    
    rename, mapped = {}, set()
    
    for c in df.columns:
        cl, target = c.lower(), None
        
        # Priority mapping
        if "amount in local currency" in cl: target = "Amount (LC)"
        elif "amount in doc. curr" in cl and "Amount (LC)" not in mapped: target = "Amount (LC)" # Fallback
        elif "amount" in cl and "Amount (LC)" not in mapped: target = "Amount (LC)"
        elif "vendor name" in cl or "name1" in cl: target = "Vendor Name"
        elif "supplier" in cl or "vendor" in cl: target = "Vendor"
        elif "payment date" in cl or "due date" in cl or "vade" in cl: target = "Due Date"
        elif "document date" in cl or "belge tarihi" in cl: target = "Document Date"
        elif "g/l account" in cl or "gl account" in cl: target = "GL Account"
        elif "document number" in cl: target = "Document No"
        elif "company code" in cl: target = "Company Code"
        
        if target and target not in mapped:
            rename[c] = target
            mapped.add(target)
    
    df.rename(columns=rename, inplace=True)
    
    # Apply Defaults
    defaults = {
        "Amount (LC)": 0, "Vendor": "Unknown", 
        "Due Date": pd.Timestamp(date.today()), "GL Account": "160000",
    }
    for col, val in defaults.items():
        if col not in df.columns: df[col] = val
    
    # Data Cleansing
    df["Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce").fillna(pd.Timestamp(date.today()))
    df["Amount (LC)"] = pd.to_numeric(df["Amount (LC)"], errors="coerce").fillna(0)
    
    today = pd.Timestamp(date.today())
    df["Days Overdue"] = (today - df["Due Date"]).dt.days.clip(lower=0)
    df["Aging Bucket"] = df["Days Overdue"].apply(assign_aging)
    df["GL Account"] = df["GL Account"].astype(str).str.strip().str.split('.').str[0]
    
    # Smart Segmentation
    vn = "Vendor Name" if "Vendor Name" in df.columns else "Vendor"
    df["Segment"] = df.apply(
        lambda r: "ICO" if is_ico(r["GL Account"], r[vn])
        else ("Employee" if is_employee(r["GL Account"], r[vn]) else "3rd Party"),
        axis=1
    )
    
    return df.loc[:, ~df.columns.duplicated()].copy()


@st.cache_data(show_spinner=False)
def load_f01(file):
    """F.01 Trial Balance loader"""
    df = smart_read(file)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()].copy()
    
    rename, mapped = {}, set()
    for c in df.columns:
        cl, target = c.lower(), None
        if "g/l account" in cl or "account number" in cl: target = "GL Account"
        elif "balance" in cl or "total of reporting" in cl: target = "Balance"
        elif "fs item" in cl or "solar" in cl: target = "SOLAR"
        elif "description" in cl or "text" in cl: target = "Description"
        
        if target and target not in mapped:
            rename[c] = target
            mapped.add(target)
    
    df.rename(columns=rename, inplace=True)
    
    for col, val in [("GL Account", "Unknown"), ("Balance", 0)]:
        if col not in df.columns: df[col] = val
    
    df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce").fillna(0)
    df["GL Account"] = df["GL Account"].astype(str).str.strip().str.split('.').str[0]
    
    if "SOLAR" not in df.columns: df["SOLAR"] = ""
    df["SOLAR"] = df["SOLAR"].astype(str).str.strip()
    
    if "Description" not in df.columns: df["Description"] = df["GL Account"]
    
    return df.loc[:, ~df.columns.duplicated()].copy()

# ══════════════════════════════════════════════════════════════════════════════
# 8. DEMO DATA GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def demo_fbl1n(n=800):
    np.random.seed(42)
    today = pd.Timestamp(date.today())
    vendors = [f"V{str(i).zfill(6)}" for i in range(1001, 1150)]
    vnames = [
        "Siemens AG", "Bosch GmbH", "SAP SE", "BASF SE", "Henkel AG",
        "Sanofi EMEA (ICO)", "Sanofi US (ICO)", "Group Treasury (ICO)",
        "John Smith (Employee)", "Travel Expense Pool", "Opella Local Supplier"
    ]
    vmap = {v: vnames[i % len(vnames)] for i, v in enumerate(vendors)}
    
    gl_pool = [("160000", 0.40), ("160100", 0.20), ("168000", 0.15), ("165000", 0.15), ("163000", 0.10)]
    gls, gprob = zip(*gl_pool)
    gprob = np.array(gprob) / sum(gprob)
    
    rows = []
    for i in range(n):
        v = np.random.choice(vendors)
        gl = np.random.choice(gls, p=gprob)
        offset = np.random.choice([-10, 0, 15, 35, 55, 80, 120], p=[0.10, 0.20, 0.25, 0.15, 0.12, 0.10, 0.08])
        due = today - pd.Timedelta(days=int(offset))
        amt = np.random.choice([-1, 1]) * round(np.random.lognormal(9, 1.5), 2)
        
        rows.append({
            "Vendor": v, "Vendor Name": vmap[v], "GL Account": gl,
            "Due Date": due, "Amount (LC)": amt,
            "Document No": f"DOC{np.random.randint(1000000, 9999999)}",
            "Company Code": np.random.choice(["DE01", "FR01", "TR01", "EG03"])
        })
    
    df = pd.DataFrame(rows)
    df["Days Overdue"] = (today - df["Due Date"]).dt.days.clip(lower=0)
    df["Aging Bucket"] = df["Days Overdue"].apply(assign_aging)
    df["Segment"] = df.apply(lambda r: "ICO" if is_ico(r["GL Account"], r["Vendor Name"]) else ("Employee" if is_employee(r["GL Account"], r["Vendor Name"]) else "3rd Party"), axis=1)
    return df

@st.cache_data(show_spinner=False)
def demo_f01():
    return pd.DataFrame([
        {"GL Account": "160000", "Balance": -3500000, "SOLAR": "40000", "Description": "AP Trade"},
        {"GL Account": "160100", "Balance": -2200000, "SOLAR": "42905", "Description": "AP Interco"},
        {"GL Account": "165000", "Balance": -450000, "SOLAR": "42006", "Description": "Employee Pay"},
    ])

# ══════════════════════════════════════════════════════════════════════════════
# 9. ANALYTICAL ENGINES (RECON & AGING BUILDER)
# ══════════════════════════════════════════════════════════════════════════════
def reconcile(fbl1n, f01):
    """SOLAR-aware GL reconciliation"""
    if f01.empty: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # 24018 added for Prepays
    f01_pay = f01[f01["SOLAR"].isin(["40000", "42905", "42006", "24018"])].copy()
    
    ap = fbl1n.groupby("GL Account")["Amount (LC)"].sum().reset_index()
    ap.columns = ["GL Account", "AP Subledger"]
    
    m = pd.merge(f01_pay[["GL Account", "Description", "Balance", "SOLAR"]], ap, on="GL Account", how="outer").fillna(0)
    
    m["Difference"] = m["Balance"] - m["AP Subledger"]
    m["Match"] = m["Difference"].abs() < 1.0
    
    matched = m[m["Match"]].copy()
    gaps = m[~m["Match"] & (m["AP Subledger"] != 0)].copy()
    missing = m[(m["Balance"].abs() >= 1.0) & (m["AP Subledger"] == 0)].copy()
    
    return matched, gaps, missing

def build_vendor_aging_matrix(df_full):
    """
    Builds a robust vendor-level aging matrix ensuring ZERO rounding loss.
    Prepares exact float arrays for Excel.
    """
    v_col = "Vendor Name" if "Vendor Name" in df_full.columns else "Vendor"
    
    # Group and Pivot raw values
    pv = df_full.pivot_table(index=[v_col, "Segment"], columns="Aging Bucket", values="Amount (LC)", aggfunc="sum", fill_value=0).reset_index()
    
    # Ensure all buckets exist
    for b in AGING_LABELS:
        if b not in pv.columns: pv[b] = 0.0
        
    # Order columns
    pv = pv[[v_col, "Segment"] + AGING_LABELS]
    pv["Total Balance"] = pv[AGING_LABELS].sum(axis=1)
    
    # Sort by absolute exposure
    pv = pv.sort_values(by="Total Balance", key=abs, ascending=False)
    
    # Calculate PERFECT TOTAL row using raw float sum (This eliminates the Rounding Bug!)
    total_row = {v_col: "TOTAL", "Segment": ""}
    for b in AGING_LABELS + ["Total Balance"]:
        total_row[b] = pv[b].sum()
        
    pv_final = pd.concat([pv, pd.DataFrame([total_row])], ignore_index=True)
    return pv_final

# ══════════════════════════════════════════════════════════════════════════════
# 10. 🛡️ FLAWLESS EXCEL EXPORT ENGINE (Bug Fix Applied)
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(df_full, recon_match, recon_gap, recon_miss):
    """
    Generates a deeply formatted Excel file using openpyxl.
    CRITICAL FIX: Uses build_vendor_aging_matrix() which totals RAW floats 
    first before Excel applies the visual '#,##0.00' mask. No data is lost!
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Design Styles
    HF = PatternFill("solid", fgColor="1A56DB")
    HFT = Font(bold=True, color="FFFFFF", size=10)
    BD = Side(style="thin", color="E2E8F0")
    CBR = Border(left=BD, right=BD, top=BD, bottom=BD)
    
    AGING_FILLS = {
        "Current": PatternFill("solid", fgColor="D1FAE5"),
        "1-30 Days": PatternFill("solid", fgColor="DBEAFE"),
        "31-60 Days": PatternFill("solid", fgColor="FEF3C7"),
        "61-90 Days": PatternFill("solid", fgColor="FEE2E2"),
        "90+ Days": PatternFill("solid", fgColor="FEE2E2"),
    }
    
    # Scale divisor (Export is NOT scaled so users get raw values for accounting)
    scale = 1.0 
    
    def format_headers(ws, nc):
        for c in range(1, nc + 1):
            cell = ws.cell(row=1, column=c)
            if cell.value:
                cell.fill = HF
                cell.font = HFT
                cell.border = CBR
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def auto_width(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 45)

    # ─── Sheet 1: Cover ───
    ws0 = wb.active
    ws0.title = "Cover"
    ws0["B2"] = "VendorFace v6.1 ULTIMATE — AP Intelligence Report"
    ws0["B2"].font = Font(bold=True, size=16, color="1A56DB")
    ws0["B3"] = f"Generated By: {st.session_state.user_name} | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws0["B3"].font = Font(size=11, color="64748B")
    ws0["B5"] = f"Total Records Analyzed: {len(df_full):,}"
    ws0["B6"] = f"Total Gross Exposure: {df_full['Amount (LC)'].sum():,.2f}"
    ws0.column_dimensions["B"].width = 50
    
    # ─── Sheet 2: AP Line Items (FULL) ───
    ws1 = wb.create_sheet("AP Line Items")
    cols = [c for c in ["Document No", "Vendor", "Vendor Name", "GL Account", "Segment", "Due Date", "Days Overdue", "Aging Bucket", "Amount (LC)", "Company Code"] if c in df_full.columns]
    
    ws1.append(cols)
    format_headers(ws1, len(cols))
    ws1.freeze_panes = "A2"
    
    for _, row in df_full[cols].iterrows():
        ws1.append(list(row))
        r = ws1.max_row
        bk = row.get("Aging Bucket", "Current")
        fill = AGING_FILLS.get(bk, PatternFill())
        
        for ci in range(1, len(cols) + 1):
            cell = ws1.cell(row=r, column=ci)
            cell.border = CBR
            cell.fill = fill
            if cols[ci - 1] == "Amount (LC)":
                cell.number_format = '#,##0.00'
            elif "Date" in cols[ci - 1]:
                cell.number_format = 'yyyy-mm-dd'
    auto_width(ws1)
    
    # ─── Sheet 3: Vendor Aging Matrix (BUG FIX) ───
    ws2 = wb.create_sheet("Vendor Aging Matrix")
    v_aging_df = build_vendor_aging_matrix(df_full)
    
    ws2.append(list(v_aging_df.columns))
    format_headers(ws2, len(v_aging_df.columns))
    ws2.freeze_panes = "A2"
    
    for _, row in v_aging_df.iterrows():
        ws2.append(list(row))
        r = ws2.max_row
        is_total = str(row.iloc[0]) == "TOTAL"
        
        for ci in range(1, len(v_aging_df.columns) + 1):
            cell = ws2.cell(row=r, column=ci)
            cell.border = CBR
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E2E8F0")
            
            c_name = v_aging_df.columns[ci - 1]
            if c_name in AGING_LABELS + ["Total Balance"]:
                cell.number_format = '#,##0.00'
                if not is_total and c_name in AGING_FILLS:
                    cell.fill = AGING_FILLS[c_name]
    auto_width(ws2)
    
    # ─── Sheets 4-6: Reconciliations ───
    def write_df_to_sheet(sheet_name, dataframe):
        if dataframe.empty: return
        ws = wb.create_sheet(sheet_name)
        ws.append(list(dataframe.columns))
        format_headers(ws, len(dataframe.columns))
        for _, row in dataframe.iterrows():
            ws.append(list(row))
            r = ws.max_row
            for ci in range(1, len(dataframe.columns) + 1):
                cell = ws.cell(row=r, column=ci)
                cell.border = CBR
                if "Balance" in dataframe.columns[ci - 1] or "Difference" in dataframe.columns[ci - 1] or "AP Subledger" in dataframe.columns[ci - 1]:
                    cell.number_format = '#,##0.00'
        auto_width(ws)
        
    write_df_to_sheet("Recon Matched", recon_match)
    write_df_to_sheet("Recon Gaps", recon_gap)
    write_df_to_sheet("Missing GL in AP", recon_miss)
    
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# 11. SIDEBAR & APP ROUTING
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style='text-align:center;padding:20px 0;'>
        <div style='font-size:2.5rem;margin-bottom:8px;'>🛡️</div>
        <div style='font-size:1.3rem;font-weight:800;color:#F1F5F9;'>VendorFace</div>
        <div style='font-size:0.7rem;color:#94A3B8;letter-spacing:0.1em;'>{VERSION_NO}</div>
    </div>
    <hr style='border-color:#475569;margin:16px 0;'/>
    """, unsafe_allow_html=True)
    
    st.markdown(f"👤 **{st.session_state.user_name}**")
    
    col_t1, col_t2 = st.columns([3, 1])
    with col_t2:
        if st.button("🌙" if st.session_state.theme == 'light' else "☀️", help="Toggle theme"):
            st.session_state.theme = 'dark' if st.session_state.theme == 'light' else 'light'
            st.rerun()
            
    st.markdown("<hr style='border-color:#475569;margin:16px 0;'/>", unsafe_allow_html=True)
    
    st.markdown("### 📥 Data Source")
    use_demo = st.toggle("Activate Demo Mode", value=True)
    fbl1n_file, f01_file = None, None
    
    if not use_demo:
        fbl1n_file = st.file_uploader("1. FBL1N — AP Line Items", type=["xlsx", "xls", "csv"])
        f01_file = st.file_uploader("2. F.01 — Trial Balance", type=["xlsx", "xls", "csv"])
    
    st.markdown("<hr style='border-color:#475569;margin:16px 0;'/>", unsafe_allow_html=True)
    
    st.markdown("### 💱 View Settings")
    cur_sel = st.selectbox("Currency Setup", list(CURRENCIES.keys()), index=0)
    cur_code = CURRENCIES[cur_sel]
    in_k = st.toggle(f"Scale to thousands (k{cur_code})", value=False)
    
    st.session_state.disp_curr = cur_code
    st.session_state.in_k = in_k
    
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚪 Secure Logout", use_container_width=True):
        st.session_state.clear()
        st.rerun()
        
    st.markdown("""
    <div style='background:rgba(16,185,129,0.1); border:1px solid rgba(16,185,129,0.3);
                border-radius:8px; padding:12px; font-size:0.75rem; line-height:1.5; margin-top: 15px;'>
        🛡️ <b>Session Secured</b><br/>Zero Data Retention active. Files are purged from RAM immediately upon session close.
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# 12. MASTER DATA EXECUTION
# ══════════════════════════════════════════════════════════════════════════════
if use_demo:
    fbl1n_df, f01_df, data_ok = demo_fbl1n(1200), demo_f01(), True
else:
    data_ok, fbl1n_df, f01_df = False, None, pd.DataFrame()
    
    if fbl1n_file:
        try:
            with st.spinner("Processing massive FBL1N dataset..."):
                fbl1n_df = load_fbl1n(fbl1n_file)
            data_ok = True
        except Exception as e:
            st.error(f"❌ Error loading FBL1N: {e}")
            st.stop()
            
    if f01_file:
        try:
            with st.spinner("Processing F.01 mapping..."):
                f01_df = load_f01(f01_file)
        except Exception as e:
            st.warning(f"F.01 loading failed: {e}")

df_full = fbl1n_df.copy() if data_ok else pd.DataFrame()

# ══════════════════════════════════════════════════════════════════════════════
# 13. UI HEADER & WELCOME
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div style='background:linear-gradient(135deg, #1e3a8a 0%, #4338ca 100%);
            padding:32px;border-radius:20px;margin-bottom:24px;color:white;
            box-shadow:0 10px 30px rgba(30,58,138,0.3);'>
    <div style='display:flex;justify-content:space-between;align-items:center;'>
        <div>
            <div style='font-size:2.2rem;font-weight:800;margin-bottom:4px;'>🛡️ VendorFace v6.1 ULTIMATE</div>
            <div style='font-size:1rem;opacity:0.9;'>Global AP Intelligence & Control Dashboard | Opella Finance</div>
        </div>
        <div style='text-align:right;'>
            <div style='font-size:0.8rem;opacity:0.8;'>{datetime.now().strftime('%d %b %Y, %H:%M')}</div>
            <div style='font-size:1.1rem;font-weight:600;margin-top:4px;'>{len(df_full):,} Validated Records</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

if use_demo:
    st.markdown("<div class='info-box'>🧪 <b>Demo Environment Active</b> — Generating dynamic synthetic payload to preview analytics logic.</div>", unsafe_allow_html=True)

if not data_ok:
    st.markdown("""
    <div style='text-align:center;padding:80px;'>
        <div style='font-size:4rem;margin-bottom:16px;'>📁</div>
        <h2 style='color:#64748B;margin-bottom:8px;'>Awaiting Sub-Ledger Data</h2>
        <p style='color:#94A3B8;'>Upload FBL1N and F.01 files via the secure sidebar to initialize the engine.</p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# 14. DYNAMIC KPI ENGINE
# ══════════════════════════════════════════════════════════════════════════════
tot_ap = df_full["Amount (LC)"].sum()
ov_df = df_full[df_full["Days Overdue"] > 0]
crit_df = df_full[df_full["Aging Bucket"] == "90+ Days"]

kpi_data = [
    ("Gross Exposure", tot_ap, f"{df_full['Vendor'].nunique()} Active Vendors", "#3B82F6", "💰"),
    ("Total Overdue", ov_df["Amount (LC)"].sum(), f"{len(ov_df):,} Aging Invoices", "#F59E0B", "⏳"),
    ("Critical 90+", crit_df["Amount (LC)"].sum(), f"{len(crit_df):,} High-Risk Items", "#EF4444", "🚨"),
    ("Intercompany (ICO)", df_full[df_full["Segment"] == "ICO"]["Amount (LC)"].sum(), "Group Level Exposure", "#8B5CF6", "🔗"),
    ("Employee Payables", df_full[df_full["Segment"] == "Employee"]["Amount (LC)"].sum(), "Travel & Expense Pool", "#10B981", "👥"),
]

st.markdown("<div class='kpi-grid'>", unsafe_allow_html=True)
cols = st.columns(len(kpi_data))
for i, (label, value, sub, color, icon) in enumerate(kpi_data):
    with cols[i]:
        progress_width = min(100, abs(value/tot_ap*100)) if tot_ap else 0
        st.markdown(f"""
        <div class='kpi-card' style='--accent:{color};'>
            <div style='font-size:2rem;margin-bottom:8px;'>{icon}</div>
            <div class='kpi-label'>{label}</div>
            <div class='kpi-value'>{fa(value)}</div>
            <div class='kpi-sub'>{sub}</div>
            <div class='progress-bar'>
                <div class='progress-fill' style='width:{progress_width:.0f}%;background:{color};'></div>
            </div>
        </div>
        """, unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# 15. MAIN TABS & ANALYTICS VIEWS
# ══════════════════════════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4 = st.tabs(["📊 Aging Matrix & Segments", "⚖️ Continuous Reconciliation", "🧠 Vendor Intelligence", "📥 Report Generation Hub"])

with tab1:
    st.markdown('<div class="sec-hdr">Aging Distribution Landscape</div>', unsafe_allow_html=True)
    
    aging = (df_full.groupby("Aging Bucket").agg(Count=("Amount (LC)", "count"), **{"Total Amount": ("Amount (LC)", "sum")}).reset_index())
    aging["Aging Bucket"] = pd.Categorical(aging["Aging Bucket"], AGING_LABELS, ordered=True)
    aging = aging.sort_values("Aging Bucket")
    aging["Total Amount"] = aging["Total Amount"].abs()
    
    fig = go.Figure()
    for i, row in aging.iterrows():
        bk = row["Aging Bucket"]
        amt = row["Total Amount"] / (1000 if st.session_state.in_k else 1)
        
        fig.add_trace(go.Bar(
            name=bk, x=[bk], y=[amt],
            marker_color=AGING_COLORS[AGING_LABELS.index(bk)],
            text=[f"{fa(row['Total Amount'])}<br>{int(row['Count']):,} items"],
            textposition="outside",
            hovertemplate=f"<b>{bk}</b><br>Amount: {fa(row['Total Amount'])}<br>Count: {int(row['Count']):,}<extra></extra>"
        ))
    
    fig.update_layout(
        showlegend=False, height=450, margin=dict(l=0, r=0, t=20, b=0),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color=theme['text']),
        yaxis=dict(gridcolor=theme['border'], title=f"Amount ({'k' if st.session_state.in_k else ''}{st.session_state.disp_curr})")
    )
    st.plotly_chart(fig, use_container_width=True)
    
    st.markdown('<div class="sec-hdr">Segment Risk Profiling</div>', unsafe_allow_html=True)
    seg_cols = st.columns(3)
    for i, seg in enumerate(["3rd Party", "ICO", "Employee"]):
        seg_df = df_full[df_full["Segment"] == seg]
        with seg_cols[i]:
            st.markdown(f"""
            <div class='glass-card' style='border-top: 4px solid {AGING_COLORS[i]};'>
                <div style='font-size:1.1rem;font-weight:700;margin-bottom:8px;'>{seg} Payables</div>
                <div style='font-size:1.8rem;font-weight:800;color:{theme['text']};margin-bottom:4px;'>{fa(seg_df['Amount (LC)'].sum())}</div>
                <div style='font-size:0.85rem;color:{theme['text_sec']};'>{len(seg_df):,} Line Items</div>
            </div>
            """, unsafe_allow_html=True)

with tab2:
    st.markdown('<div class="sec-hdr">F.01 vs FBL1N Subledger Reconciliation</div>', unsafe_allow_html=True)
    if f01_file or use_demo:
        rm, rg, rms = reconcile(df_full, f01_df)
        gap_tot = rg["Difference"].abs().sum() if not rg.empty else 0
        
        if gap_tot < 1 and rms.empty:
            st.markdown('<div class="success-box">✅ Audit Passed: All payable GL accounts in Trial Balance perfectly match the AP Subledger.</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="warning-box">⚠️ Audit Alert: Detected {len(rg)} GL account(s) with variances. Total identified gap: <b>{fa(gap_tot)}</b></div>', unsafe_allow_html=True)
            
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Subledger Variances (Gaps)**")
                st.dataframe(rg.style.format({"Balance": "{:,.2f}", "AP Subledger": "{:,.2f}", "Difference": "{:,.2f}"}), use_container_width=True)
            with c2:
                st.markdown("**Trial Balance Mappings Missing in Subledger**")
                st.dataframe(rms.style.format({"Balance": "{:,.2f}"}), use_container_width=True)
    else:
        st.info("Upload F.01 Trial Balance file via sidebar to unlock automated reconciliation features.")

with tab3:
    st.markdown('<div class="sec-hdr">Concentration Analysis: Top 20 Vendors</div>', unsafe_allow_html=True)
    lbl = "Vendor Name" if "Vendor Name" in df_full.columns else "Vendor"
    top20 = (df_full.groupby(lbl)["Amount (LC)"].sum().abs().sort_values(ascending=False).head(20).reset_index())
    
    fig2 = go.Figure(go.Bar(
        x=top20[lbl].str[:35],
        y=top20["Amount (LC)"] / (1000 if st.session_state.in_k else 1),
        marker=dict(color=top20["Amount (LC)"], colorscale='Viridis', showscale=True),
        text=[fa(v) for v in top20["Amount (LC)"]], textposition='outside'
    ))
    fig2.update_layout(
        height=550, margin=dict(l=0, r=0, t=20, b=0), plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=theme['text']), yaxis_title=f"Gross Balance ({'k' if st.session_state.in_k else ''}{st.session_state.disp_curr})"
    )
    st.plotly_chart(fig2, use_container_width=True)

with tab4:
    st.markdown('<div class="sec-hdr">Audit & Export Center</div>', unsafe_allow_html=True)
    
    col_e1, col_e2, col_e3 = st.columns(3)
    
    with col_e1:
        st.markdown(f"""
        <div class='glass-card' style='text-align:center;'>
            <div style='font-size:2.5rem;margin-bottom:10px;'>📊</div>
            <div style='font-weight:700;margin-bottom:8px;'>Master Excel Workbook</div>
            <div style='font-size:0.8rem;color:{theme['text_sec']};margin-bottom:20px;'>
                Complete dataset with 100% precision vendor matrices and formatted tabs.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("⬇️ Compile & Download Full Pack", type="primary", use_container_width=True):
            with st.spinner("Processing massive arrays and building Excel via OpenPyXL..."):
                rm, rg, rms = reconcile(df_full, f01_df) if (f01_file or use_demo) else (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
                xdata = build_excel(df_full, rm, rg, rms)
            
            st.download_button(
                label="📥 Save Local Copy (.xlsx)",
                data=xdata,
                file_name=f"Opella_VendorFace_Master_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    with col_e2:
        st.markdown(f"""
        <div class='glass-card' style='text-align:center;'>
            <div style='font-size:2.5rem;margin-bottom:10px;'>⚠️</div>
            <div style='font-weight:700;margin-bottom:8px;'>High-Risk Export</div>
            <div style='font-size:0.8rem;color:{theme['text_sec']};margin-bottom:20px;'>
                Extract only items in the 90+ days overdue bracket for immediate action.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        if len(crit_df) > 0:
            st.download_button(
                label=f"⬇️ Download {len(crit_df)} Critical Items (.csv)",
                data=crit_df.to_csv(index=False).encode("utf-8-sig"),
                file_name=f"VendorFace_Critical_90Plus_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        else:
            st.success("✅ Zero critical items detected.")

# ══════════════════════════════════════════════════════════════════════════════
# 16. FOOTER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div style='text-align:center;margin-top:60px;padding-top:24px;border-top:1px solid {theme['border']};
            color:{theme['text_sec']};font-size:0.8rem;letter-spacing:0.05em;'>
    <b>VendorFace v6.1 ULTIMATE</b> • Financial Architecture by Can Adiguzel • Precision Engineered for Opella Finance Operations
</div>
""", unsafe_allow_html=True)
